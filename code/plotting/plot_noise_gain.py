from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import matplotlib as mpl
mpl.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


NOISE_ORDER = ["no_noise", "SNR=20", "SNR=15", "SNR=10", "SNR=5", "SNR=0"]
BASELINE_NAME = "Basic Model"
COMPARISON_SERIES = ["Model Denoising", "Wavelet Denoising", "Hybrid Denoising"]

SERIES_STYLES = {
    "Model Denoising": {"color": "#9CC3E6", "hatch": "//"},
    "Wavelet Denoising": {"color": "#4F86C6", "hatch": "xx"},
    "Hybrid Denoising": {"color": "#8B1E3F", "hatch": ".."},
}
OUTPUT_FORMATS = ("pdf", "png", "svg")


def setup_plot_style() -> None:
    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": ["Times New Roman", "DejaVu Serif"],
            "axes.labelsize": 12,
            "axes.titlesize": 13,
            "legend.fontsize": 10,
            "xtick.labelsize": 10,
            "ytick.labelsize": 10,
            "axes.linewidth": 1.0,
            "axes.unicode_minus": False,
            "figure.dpi": 150,
            "savefig.dpi": 600,
        }
    )


def normalize_noise_label(label: object) -> str:
    if label is None:
        return ""
    text = str(label).strip().replace(" ", "")
    if not text:
        return ""
    lower = text.lower()
    if lower in {"no_noise", "nonoise", "clean", "normal"}:
        return "no_noise"
    if lower.startswith("snr"):
        number = "".join(ch for ch in text if ch.isdigit() or ch == "-")
        return f"SNR={number}" if number else text
    return text


def read_table_from_workbook(path: Path) -> Tuple[str, Dict[str, Dict[str, float]]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    series_names = [str(value).strip() for value in header[1:] if value is not None and str(value).strip()]
    if not series_names:
        raise ValueError(f"No series names found in {path.name}")

    data: Dict[str, Dict[str, float]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(value is None for value in row):
            continue
        noise_label = normalize_noise_label(row[0])
        if not noise_label:
            continue
        values: Dict[str, float] = {}
        for series_name, value in zip(series_names, row[1:]):
            if value is None:
                continue
            values[series_name] = float(value)
        if values:
            data[noise_label] = values

    dataset_name = sheet.title.strip() if sheet.title else path.stem
    return dataset_name, data


def get_noise_order_for_data(data: Dict[str, Dict[str, float]]) -> List[str]:
    return [noise for noise in NOISE_ORDER if noise in data]


def build_delta_arrays(data: Dict[str, Dict[str, float]], noise_order: Sequence[str]) -> Dict[str, np.ndarray]:
    baseline = np.array([data.get(noise, {}).get(BASELINE_NAME, np.nan) for noise in noise_order], dtype=float)
    deltas: Dict[str, np.ndarray] = {}
    for series_name in COMPARISON_SERIES:
        values = np.array([data.get(noise, {}).get(series_name, np.nan) for noise in noise_order], dtype=float)
        deltas[series_name] = values - baseline
    deltas[BASELINE_NAME] = baseline
    return deltas


def add_sign_guides(ax: plt.Axes) -> None:
    ax.axhline(0.0, color="#444444", linewidth=1.0)
    ax.axvline(0.0, color="#B0B0B0", linewidth=0.8, linestyle="--")
    ax.grid(True, axis="y", linestyle="--", linewidth=0.8, alpha=0.35)


def save_figure(fig: plt.Figure, output_dir: Path, stem: str) -> None:
    for output_format in OUTPUT_FORMATS:
        fig.savefig(output_dir / f"{stem}.{output_format}", bbox_inches="tight")


def plot_dataset_gain(dataset_name: str, data: Dict[str, Dict[str, float]], output_dir: Path) -> None:
    noise_order = get_noise_order_for_data(data)
    deltas = build_delta_arrays(data, noise_order)
    x = np.arange(len(noise_order), dtype=float)
    width = 0.24
    offsets = [-width, 0.0, width]

    baseline = deltas[BASELINE_NAME]
    model = deltas["Model Denoising"]
    wavelet = deltas["Wavelet Denoising"]
    hybrid = deltas["Hybrid Denoising"]
    interaction = hybrid - (model + wavelet)

    fig, axes = plt.subplots(1, 2, figsize=(13.6, 5.2), constrained_layout=True)

    ax = axes[0]
    for index, series_name in enumerate(COMPARISON_SERIES):
        values = deltas[series_name]
        style = SERIES_STYLES[series_name]
        ax.bar(
            x + offsets[index],
            values,
            width=width,
            label=series_name,
            color=style["color"],
            edgecolor="#333333",
            linewidth=0.7,
            hatch=style["hatch"],
        )

    add_sign_guides(ax)
    ax.set_xticks(x)
    ax.set_xticklabels(noise_order)
    ax.set_xlabel("Noise Condition")
    ax.set_ylabel(r"$\Delta$Accuracy")
    ax.set_title("Relative gain over Basic Model")
    max_abs = np.nanmax(np.abs(np.concatenate([model, wavelet, hybrid]))) if np.isfinite(np.concatenate([model, wavelet, hybrid])).any() else 0.1
    ax.set_ylim(-max(0.08, max_abs * 1.25), max(0.08, max_abs * 1.25))
    ax.set_xlim(-0.6, len(noise_order) - 0.4)
    ax.legend(loc="upper left", frameon=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(direction="in", length=4, width=1)

    ax2 = axes[1]
    ax2.bar(x, interaction, width=0.58, color="#8B1E3F", edgecolor="#333333", linewidth=0.7)
    add_sign_guides(ax2)
    ax2.set_xticks(x)
    ax2.set_xticklabels(noise_order)
    ax2.set_xlabel("Noise Condition")
    ax2.set_ylabel(r"$\Delta$Accuracy")
    ax2.set_title("Hybrid Interaction Effect")
    max_interaction = np.nanmax(np.abs(interaction)) if np.isfinite(interaction).any() else 0.1
    ax2.set_ylim(-max(0.08, max_interaction * 1.25), max(0.08, max_interaction * 1.25))
    ax2.set_xlim(-0.6, len(noise_order) - 0.4)
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_visible(False)
    ax2.tick_params(direction="in", length=4, width=1)

    fig.suptitle(dataset_name, y=1.02, fontsize=13)
    base_name = dataset_name or "dataset"
    save_figure(fig, output_dir, f"{base_name}_gain")
    plt.close(fig)


def plot_combined(dataset_items: Sequence[Tuple[str, Dict[str, Dict[str, float]]]], output_dir: Path) -> None:
    fig, axes = plt.subplots(len(dataset_items), 2, figsize=(13.6, 9.0))
    if len(dataset_items) == 1:
        axes = np.array([axes])

    for row_index, (dataset_name, data) in enumerate(dataset_items):
        noise_order = get_noise_order_for_data(data)
        x = np.arange(len(noise_order), dtype=float)
        width = 0.24
        offsets = [-width, 0.0, width]
        deltas = build_delta_arrays(data, noise_order)
        model = deltas["Model Denoising"]
        wavelet = deltas["Wavelet Denoising"]
        hybrid = deltas["Hybrid Denoising"]
        interaction = hybrid - (model + wavelet)

        ax = axes[row_index, 0]
        for index, series_name in enumerate(COMPARISON_SERIES):
            values = deltas[series_name]
            style = SERIES_STYLES[series_name]
            ax.bar(
                x + offsets[index],
                values,
                width=width,
                label=series_name if row_index == 0 else None,
                color=style["color"],
                edgecolor="#333333",
                linewidth=0.7,
                hatch=style["hatch"],
            )
        add_sign_guides(ax)
        ax.set_xticks(x)
        ax.set_xticklabels(noise_order)
        ax.set_ylabel(r"$\Delta$Accuracy")
        ax.set_title(f"{dataset_name} | gain over Basic Model")
        max_abs = np.nanmax(np.abs(np.concatenate([model, wavelet, hybrid]))) if np.isfinite(np.concatenate([model, wavelet, hybrid])).any() else 0.1
        ax.set_ylim(-max(0.08, max_abs * 1.25), max(0.08, max_abs * 1.25))
        ax.set_xlim(-0.6, len(noise_order) - 0.4)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(direction="in", length=4, width=1)
        ax.text(0.02, 0.96, f"({chr(97 + row_index * 2)})", transform=ax.transAxes, ha="left", va="top", fontsize=11)

        ax2 = axes[row_index, 1]
        ax2.bar(x, interaction, width=0.58, color="#8B1E3F", edgecolor="#333333", linewidth=0.7)
        add_sign_guides(ax2)
        ax2.set_xticks(x)
        ax2.set_xticklabels(noise_order)
        ax2.set_ylabel(r"$\Delta$Accuracy")
        ax2.set_title(f"{dataset_name} | hybrid interaction effect")
        max_interaction = np.nanmax(np.abs(interaction)) if np.isfinite(interaction).any() else 0.1
        ax2.set_ylim(-max(0.08, max_interaction * 1.25), max(0.08, max_interaction * 1.25))
        ax2.set_xlim(-0.6, len(noise_order) - 0.4)
        ax2.spines["top"].set_visible(False)
        ax2.spines["right"].set_visible(False)
        ax2.tick_params(direction="in", length=4, width=1)
        ax2.text(0.02, 0.96, f"({chr(98 + row_index * 2)})", transform=ax2.transAxes, ha="left", va="top", fontsize=11)

    handles, labels = axes[0, 0].get_legend_handles_labels()
    fig.suptitle("Overall gain and interaction effect under noise", y=0.985, fontsize=13)
    fig.legend(handles, labels, loc="upper center", ncol=3, frameon=False, bbox_to_anchor=(0.5, 0.955))
    fig.subplots_adjust(top=0.90)
    save_figure(fig, output_dir, "combined_gain")
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description="Plot gain charts relative to Basic Model for paper figures.")
    parser.add_argument("--input-dir", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent / "gain_figures")
    parser.add_argument(
        "--workbooks",
        nargs="*",
        default=["试验结果CWRU.xlsx", "试验结果SUFD.xlsx"],
        help="Workbook filenames located under --input-dir.",
    )
    parser.add_argument("--mode", choices=["single", "combined", "both"], default="both")
    args = parser.parse_args()

    setup_plot_style()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    dataset_items: List[Tuple[str, Dict[str, Dict[str, float]]]] = []
    for workbook_name in args.workbooks:
        workbook_path = args.input_dir / workbook_name
        dataset_name, data = read_table_from_workbook(workbook_path)
        dataset_items.append((dataset_name, data))

        if args.mode in {"single", "both"}:
            plot_dataset_gain(dataset_name, data, args.output_dir)

    if args.mode in {"combined", "both"}:
        plot_combined(dataset_items, args.output_dir)

    print(f"Saved gain figures to: {args.output_dir}")


if __name__ == "__main__":
    main()
