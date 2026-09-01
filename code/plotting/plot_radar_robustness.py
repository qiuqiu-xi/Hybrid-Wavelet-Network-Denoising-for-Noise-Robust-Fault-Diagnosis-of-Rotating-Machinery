from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import matplotlib as mpl
mpl.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


NOISE_ORDER = ["no_noise", "SNR=20", "SNR=15", "SNR=10", "SNR=5", "SNR=0"]
SERIES_ORDER = ["Basic Model", "Model Denoising", "Wavelet Denoising", "Hybrid Denoising"]

SERIES_STYLES = {
    "Basic Model": {"color": "#8E8E8E", "marker": "o", "linestyle": "--"},
    "Model Denoising": {"color": "#A7C7E7", "marker": "s", "linestyle": "-"},
    "Wavelet Denoising": {"color": "#4F86C6", "marker": "^", "linestyle": "-"},
    "Hybrid Denoising": {"color": "#8B1E3F", "marker": "D", "linestyle": "-"},
}
OUTPUT_FORMATS = ("pdf", "png", "svg")


def setup_plot_style() -> None:
    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": ["Times New Roman", "DejaVu Serif"],
            "axes.labelsize": 12,
            "axes.titlesize": 13,
            "legend.fontsize": 9,
            "xtick.labelsize": 10,
            "ytick.labelsize": 9,
            "axes.linewidth": 1.0,
            "axes.unicode_minus": False,
            "figure.dpi": 150,
            "savefig.dpi": 600,
        }
    )


def normalize_noise_label(label: object) -> str:
    if label is None:
        return ""
    text = str(label).strip().replace(" ", "")
    if not text:
        return ""
    lower = text.lower()
    if lower in {"no_noise", "nonoise", "clean", "normal"}:
        return "no_noise"
    if lower.startswith("snr"):
        number = "".join(ch for ch in text if ch.isdigit() or ch == "-")
        return f"SNR={number}" if number else text
    return text


def read_table_from_workbook(path: Path) -> Tuple[str, Dict[str, Dict[str, float]]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    series_names = [str(value).strip() for value in header[1:] if value is not None and str(value).strip()]
    if not series_names:
        raise ValueError(f"No series names found in {path.name}")

    data: Dict[str, Dict[str, float]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(value is None for value in row):
            continue
        noise_label = normalize_noise_label(row[0])
        if not noise_label:
            continue
        values: Dict[str, float] = {}
        for series_name, value in zip(series_names, row[1:]):
            if value is None:
                continue
            values[series_name] = float(value)
        if values:
            data[noise_label] = values

    dataset_name = sheet.title.strip() if sheet.title else path.stem
    return dataset_name, data


def get_noise_order_for_data(data: Dict[str, Dict[str, float]]) -> List[str]:
    return [noise for noise in NOISE_ORDER if noise in data]


def align_data(data: Dict[str, Dict[str, float]], noise_order: Sequence[str]) -> Dict[str, np.ndarray]:
    aligned: Dict[str, np.ndarray] = {}
    for series_name in SERIES_ORDER:
        aligned[series_name] = np.array([data.get(noise, {}).get(series_name, np.nan) for noise in noise_order], dtype=float)
    return aligned


def close_polygon(values: np.ndarray) -> np.ndarray:
    return np.concatenate([values, values[:1]])


def make_angles(count: int) -> np.ndarray:
    angles = np.linspace(0, 2 * np.pi, count, endpoint=False)
    return np.concatenate([angles, angles[:1]])


def save_figure(fig: plt.Figure, output_dir: Path, stem: str) -> None:
    for output_format in OUTPUT_FORMATS:
        fig.savefig(output_dir / f"{stem}.{output_format}", bbox_inches="tight")


def plot_dataset_radar(ax: plt.Axes, dataset_name: str, aligned: Dict[str, np.ndarray], noise_order: Sequence[str]) -> None:
    angles = make_angles(len(noise_order))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)

    for series_name in SERIES_ORDER:
        values = close_polygon(aligned[series_name])
        style = SERIES_STYLES[series_name]
        ax.plot(
            angles,
            values,
            color=style["color"],
            linewidth=2.3 if series_name == "Hybrid Denoising" else 1.9,
            linestyle=style["linestyle"],
            marker=style["marker"],
            markersize=4.5,
            markerfacecolor="white",
            markeredgewidth=0.9,
            label=series_name,
        )
        ax.fill(angles, values, color=style["color"], alpha=0.08 if series_name != "Hybrid Denoising" else 0.14)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(["No noise" if noise == "no_noise" else noise.split("=")[1] for noise in noise_order])
    ax.set_ylim(0.0, 1.05)
    ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
    ax.set_yticklabels(["0.2", "0.4", "0.6", "0.8", "1.0"])
    ax.set_rlabel_position(180)
    ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.35)
    ax.set_title(dataset_name, pad=16)


def plot_combined(dataset_items: Sequence[Tuple[str, Dict[str, np.ndarray], Sequence[str]]], output_dir: Path) -> None:
    fig, axes = plt.subplots(1, len(dataset_items), figsize=(13.8, 6.0), subplot_kw={"projection": "polar"})
    if len(dataset_items) == 1:
        axes = [axes]

    for ax, (dataset_name, aligned, noise_order) in zip(axes, dataset_items):
        plot_dataset_radar(ax, dataset_name, aligned, noise_order)

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", ncol=4, frameon=False, bbox_to_anchor=(0.5, 1.02))
    fig.suptitle("Overall robustness radar chart", y=1.06, fontsize=13)
    fig.subplots_adjust(top=0.84, bottom=0.05, left=0.04, right=0.96, wspace=0.22)
    save_figure(fig, output_dir, "combined_radar")
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description="Plot radar charts for overall robustness across noise levels.")
    parser.add_argument("--input-dir", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent / "radar_figures")
    parser.add_argument(
        "--workbooks",
        nargs="*",
        default=["试验结果CWRU.xlsx", "试验结果SUFD.xlsx"],
        help="Workbook filenames located under --input-dir.",
    )
    args = parser.parse_args()

    setup_plot_style()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    dataset_items: List[Tuple[str, Dict[str, np.ndarray], Sequence[str]]] = []
    for workbook_name in args.workbooks:
        workbook_path = args.input_dir / workbook_name
        dataset_name, raw_data = read_table_from_workbook(workbook_path)
        noise_order = get_noise_order_for_data(raw_data)
        dataset_items.append((dataset_name, align_data(raw_data, noise_order), noise_order))

    plot_combined(dataset_items, args.output_dir)
    print(f"Saved radar figures to: {args.output_dir}")


if __name__ == "__main__":
    main()
