from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import matplotlib as mpl
mpl.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


NOISE_ORDER = ["no_noise", "SNR=20", "SNR=15", "SNR=10", "SNR=5", "SNR=0"]
SERIES_ORDER = ["Basic Model", "Model Denoising", "Wavelet Denoising", "Hybrid Denoising"]

SERIES_STYLES = {
    "Basic Model": {"color": "#C7C7C7", "marker": "o", "hatch": ""},
    "Model Denoising": {"color": "#9CC3E6", "marker": "s", "hatch": "//"},
    "Wavelet Denoising": {"color": "#4F86C6", "marker": "^", "hatch": "xx"},
    "Hybrid Denoising": {"color": "#8B1E3F", "marker": "D", "hatch": ".."},
}
OUTPUT_FORMATS = ("pdf", "png", "svg")


def setup_plot_style() -> None:
    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": ["Times New Roman", "DejaVu Serif"],
            "axes.labelsize": 12,
            "axes.titlesize": 13,
            "legend.fontsize": 10,
            "xtick.labelsize": 10,
            "ytick.labelsize": 10,
            "axes.linewidth": 1.0,
            "axes.unicode_minus": False,
            "figure.dpi": 150,
            "savefig.dpi": 600,
        }
    )


def normalize_noise_label(label: object) -> str:
    if label is None:
        return ""
    text = str(label).strip().replace(" ", "")
    if not text:
        return ""
    lower = text.lower()
    if lower in {"no_noise", "nonoise", "clean", "normal"}:
        return "no_noise"
    if lower.startswith("snr"):
        number = "".join(ch for ch in text if ch.isdigit() or ch == "-")
        return f"SNR={number}" if number else text
    return text


def read_table_from_workbook(path: Path) -> Tuple[str, Dict[str, Dict[str, float]]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    series_names = [str(value).strip() for value in header[1:] if value is not None and str(value).strip()]
    if not series_names:
        raise ValueError(f"No series names found in {path.name}")

    data: Dict[str, Dict[str, float]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(value is None for value in row):
            continue
        noise_label = normalize_noise_label(row[0])
        if not noise_label:
            continue
        values: Dict[str, float] = {}
        for series_name, value in zip(series_names, row[1:]):
            if value is None:
                continue
            values[series_name] = float(value)
        if values:
            data[noise_label] = values

    dataset_name = sheet.title.strip() if sheet.title else path.stem
    return dataset_name, data


def get_noise_order_for_data(data: Dict[str, Dict[str, float]]) -> List[str]:
    return [noise for noise in NOISE_ORDER if noise in data]


def align_data(data: Dict[str, Dict[str, float]], noise_order: Sequence[str]) -> Dict[str, np.ndarray]:
    aligned: Dict[str, np.ndarray] = {}
    for series_name in SERIES_ORDER:
        aligned[series_name] = np.array([data.get(noise, {}).get(series_name, np.nan) for noise in noise_order], dtype=float)
    return aligned


def save_figure(fig: plt.Figure, output_dir: Path, stem: str) -> None:
    for output_format in OUTPUT_FORMATS:
        fig.savefig(output_dir / f"{stem}.{output_format}", bbox_inches="tight")


def plot_grouped_bar(dataset_name: str, aligned: Dict[str, np.ndarray], noise_order: Sequence[str], output_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(7.0, 4.8))
    x = np.arange(len(noise_order), dtype=float)
    width = 0.19
    offsets = (np.arange(len(SERIES_ORDER)) - (len(SERIES_ORDER) - 1) / 2) * width

    for offset, series_name in zip(offsets, SERIES_ORDER):
        style = SERIES_STYLES[series_name]
        ax.bar(
            x + offset,
            aligned[series_name],
            width=width,
            label=series_name,
            color=style["color"],
            edgecolor="#333333",
            linewidth=0.7,
            hatch=style["hatch"],
        )

    ax.set_xticks(x)
    ax.set_xticklabels(noise_order)
    ax.set_xlabel("Noise Condition")
    ax.set_ylabel("Accuracy")
    ax.set_ylim(0.0, 1.02)
    ax.set_xlim(-0.55, len(noise_order) - 0.45)
    ax.grid(True, axis="y", linestyle="--", linewidth=0.8, alpha=0.35)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(direction="in", length=4, width=1)
    ax.legend(loc="upper center", ncol=2, frameon=False, bbox_to_anchor=(0.5, 1.18))
    fig.tight_layout()

    save_figure(fig, output_dir, f"{dataset_name or 'dataset'}_accuracy_bar")
    plt.close(fig)


def plot_single_dataset(dataset_name: str, aligned: Dict[str, np.ndarray], noise_order: Sequence[str], output_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(7.0, 4.8))
    x = np.arange(len(noise_order))

    for series_name in SERIES_ORDER:
        y = aligned[series_name]
        style = SERIES_STYLES[series_name]
        ax.plot(
            x,
            y,
            label=series_name,
            color=style["color"],
            marker=style["marker"],
            linewidth=2.2,
            markersize=6.0,
            markerfacecolor="white",
            markeredgewidth=1.1,
        )

    ax.set_xticks(x)
    ax.set_xticklabels(noise_order)
    ax.set_xlabel("Noise Condition")
    ax.set_ylabel("Accuracy")
    ax.set_ylim(0.0, 1.02)
    ax.set_xlim(-0.2, len(noise_order) - 0.8)
    ax.grid(True, axis="y", linestyle="--", linewidth=0.8, alpha=0.35)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(direction="in", length=4, width=1)
    ax.legend(loc="upper center", ncol=2, frameon=False, bbox_to_anchor=(0.5, 1.18))
    fig.tight_layout()

    base_name = dataset_name or "dataset"
    save_figure(fig, output_dir, f"{base_name}_noise_accuracy")
    plt.close(fig)


def plot_combined(dataset_items: Sequence[Tuple[str, Dict[str, np.ndarray], Sequence[str]]], output_dir: Path) -> None:
    fig, axes = plt.subplots(1, len(dataset_items), figsize=(13.0, 5.0), sharey=True, constrained_layout=True)
    if len(dataset_items) == 1:
        axes = [axes]

    for index, ((dataset_name, aligned, noise_order), ax) in enumerate(zip(dataset_items, axes), start=1):
        x = np.arange(len(noise_order))
        for series_name in SERIES_ORDER:
            y = aligned[series_name]
            style = SERIES_STYLES[series_name]
            ax.plot(
                x,
                y,
                label=series_name,
                color=style["color"],
                marker=style["marker"],
                linewidth=2.2,
                markersize=5.6,
                markerfacecolor="white",
                markeredgewidth=1.0,
            )

        ax.set_xticks(x)
        ax.set_xticklabels(noise_order, rotation=0)
        ax.set_title(dataset_name)
        ax.set_xlabel("Noise Condition")
        ax.set_ylim(0.0, 1.02)
        ax.set_xlim(-0.2, len(noise_order) - 0.8)
        ax.grid(True, axis="y", linestyle="--", linewidth=0.8, alpha=0.35)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(direction="in", length=4, width=1)
        ax.text(0.02, 0.96, f"({chr(96 + index)})", transform=ax.transAxes, ha="left", va="top", fontsize=11)

    axes[0].set_ylabel("Accuracy")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", ncol=4, frameon=False, bbox_to_anchor=(0.5, 1.04))
    save_figure(fig, output_dir, "combined_noise_accuracy")
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description="Plot noise-robust accuracy curves for paper figures.")
    parser.add_argument("--input-dir", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent / "figures")
    parser.add_argument(
        "--workbooks",
        nargs="*",
        default=["试验结果CWRU.xlsx", "试验结果SUFD.xlsx"],
        help="Workbook filenames located under --input-dir.",
    )
    parser.add_argument("--mode", choices=["single", "combined", "both"], default="both")
    args = parser.parse_args()

    setup_plot_style()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    dataset_items: List[Tuple[str, Dict[str, np.ndarray], Sequence[str]]] = []
    for workbook_name in args.workbooks:
        workbook_path = args.input_dir / workbook_name
        dataset_name, raw_data = read_table_from_workbook(workbook_path)
        noise_order = get_noise_order_for_data(raw_data)
        aligned = align_data(raw_data, noise_order)
        dataset_items.append((dataset_name, aligned, noise_order))

        if args.mode in {"single", "both"}:
            plot_grouped_bar(dataset_name, aligned, noise_order, args.output_dir)
            plot_single_dataset(dataset_name, aligned, noise_order, args.output_dir)

    if args.mode in {"combined", "both"}:
        plot_combined(dataset_items, args.output_dir)

    print(f"Saved figures to: {args.output_dir}")


if __name__ == "__main__":
    main()
